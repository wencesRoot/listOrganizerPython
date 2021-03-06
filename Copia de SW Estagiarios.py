from Tkinter import *
from openpyxl import load_workbook
from tkMessageBox import *
from openpyxl.styles import PatternFill
import warnings
import time
import os

root = Tk()
root.title("Copia de SW v0.7 - Podrik")
root.iconbitmap("CopiaIcon.ico")
root.resizable(width=FALSE, height=FALSE)
warnings.filterwarnings("ignore")

redFill = PatternFill(start_color='FFEE1111', end_color='FFEE1111', fill_type='solid')
orangeFill = PatternFill(start_color='FF6B00', end_color='FF6B00', fill_type='solid')
blueFill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')


text = StringVar()
text.set('Status')
warnings.simplefilter("ignore")

#Montando o GUI:
label1 = Label(root, text="POOLS")
label2 = Label(root, text="SWS")
label3 = Label(root,borderwidth=4, pady=5,relief= RIDGE, textvariable=text, bg="red",width=75)

label1.grid(row=1, column=0, sticky=E)
label2.grid(row=2, column=0, sticky=E)
label3.grid(row=3, columnspan=2)

entry = Entry(root,width=75)
entry2 = Entry(root,width=75)

entry.grid(row=1, column=1)
entry2.grid(row=2, column=1)

bollP=0
bollS=0
#Declara Listas vazias
pool=[]
sws=[]
sws2=[]
pool2=[]
listaP=[]
listaS=[]
listaS2=[]
listaP3=[]
listaP2=[]

#Funcoes dos botoes:
def callRules():
    showinfo("INSTRUCOES", "1-Abra um Excell novo em branco\n2-Copie e cole os Pools e os SWs do Planning nesse Excell\n3-Organize os SWs para nao deixar nenhum espaco em branco entre eles\n4-Copie e cole os Pools na parte de POOLs do programa\n5-Copie e cole os SWs na parte SWs do programa\n6-Aperte no botao POOLS\n7-Aperte no Botao SWS\n8-Aperte no Botao Excell\n9-A planilha vai ser criada na pasta do programa\n10-Fim")

def callPool():
    pool = entry.get() #Pega o pool do user
    pool2= pool.splitlines(0) #Separa em lista
    myset = set(pool2) #Tira os iguais da lista
    global listaP
    listaP = list(myset) #Transforma em lista de novo
    listaP.sort() #Organiza em ordem alfabetica
    global text
    text.set('DONE POOL')
    button2['state'] = NORMAL
    

def callSW():
    sws = entry2.get() #Pega os SWS do user
    sws2= sws.splitlines(0) #Separa em lista
    myset2 = set(sws2) #Tira os iguais da lista
    global listaS
    listaS = list(myset2) #Transforma em lista de novo
    listaS.sort() #Organiza em ordem alfabetica
    listaS.insert(0,"SWS")

    #organiza lista final de pool
    mysetP = set(listaP) - set(listaS) #Formula Padrao
    listaP2 = list(mysetP)
    global listaP3
    listaP3 = listaP2
    listaP3.sort()
    listaP3.insert(0,"POOLS")
    #listaP3.append("----Acaba Pool (Apagar isso)----")

    #organiza a lista que repete em pool e sw
    mysetos2 = set(listaP) & set(listaS)
    global listaI
    listaI = list(mysetos2)
    listaI.insert(0,"Repetidos")
    
    #junta lista final de pool e de sw
    #listaP3.extend(listaS)
    global text
    text.set('DONE SWSs')
    button3['state'] = NORMAL

def callExcel():

    wb = load_workbook('Copia de SW.xlsx')

    #Pega a aba Projetos(ws)
    ws3 = wb.get_sheet_by_name("Projetos")
    #Printa na coluna da aba, pool depois sws, depois repetidos
    for i in range(len(listaP3)):
        ws3.cell(row=i+1, column=3).value=listaP3[i]
        ws3.cell(row=i+1, column=3).fill = redFill

    for i in range(len(listaS)):
        ws3.cell(row=i+1, column=5).value=listaS[i]
        ws3.cell(row=i+1, column=5).fill = orangeFill
    
    for i in range(len(listaI)):
        ws3.cell(row=i+1, column=7).value=listaI[i]
        ws3.cell(row=i+1, column=7).fill = blueFill

    #seta o dia e o mes
    data=time.strftime("%d.%m")

    #Salva o Excel com a data de hoje
    global text
    dest_filename = 'Copia de SW %s.xlsx' %data 
    wb.save(filename = dest_filename)
    text.set("!! PLANILHA CRIADA NA PASTA DO PROGRAMA !! ")


#monta botoes no GUI
button = Button(root, text="1-POOL!", width=10, command=callPool)
button2 = Button(root, text="2-SW!", width=10, command=callSW,state=DISABLED)
button3 = Button(root, text="3-EXCEL!", width=10, command=callExcel,state=DISABLED)
button4 = Button(root,fg="ORANGE", text="!! INSTRUCOES !!",width=20, command=callRules)
button.grid(row=5, columnspan=2)
button2.grid(row=6, columnspan=2)
button3.grid(row=7, columnspan=2)
button4.grid(row=4, columnspan=2)

mainloop()
